from __future__ import annotations

import csv
import json
import os
import shutil
import tempfile
from datetime import UTC, datetime
from pathlib import Path

import yaml
from openpyxl import load_workbook

from pititino.config import Settings
from pititino.errors import ToolExecutionError, UnsupportedFileTypeError
from pititino.tools.excel import apply_change_operation
from pititino.transactions.audit import record_transaction
from pititino.transactions.changeset import ChangeSet
from pititino.workspace import Workspace


def apply_changeset(changeset: ChangeSet, workspace: Workspace, settings: Settings) -> Path:
    """Apply an approved changeset through a validated temporary file."""
    target = workspace.resolve(changeset.target, must_exist=True)
    if changeset.source_revision is not None:
        current_revision = workspace.revision(target)
        if current_revision != changeset.source_revision:
            record_transaction(
                workspace,
                changeset,
                "conflict",
                error="target changed since proposal",
            )
            raise ToolExecutionError(
                f"Target changed since the proposal was created: {changeset.target}"
            )

    backup_path: Path | None = None
    temp_path: Path | None = None
    try:
        with tempfile.NamedTemporaryFile(
            prefix=f".{target.stem}.", suffix=target.suffix, dir=target.parent, delete=False
        ) as temporary:
            temp_path = Path(temporary.name)
        restore_operation = next(
            (operation for operation in changeset.operations if operation.operation == "restore_backup"),
            None,
        )
        if restore_operation is not None and len(changeset.operations) != 1:
            raise ToolExecutionError("Backup restore cannot be combined with other operations")
        if restore_operation is not None:
            source = _resolve_backup(workspace, restore_operation.arguments["backup"])
            shutil.copy2(source, temp_path)
        else:
            shutil.copy2(target, temp_path)
        if restore_operation is not None:
            pass
        elif target.suffix.lower() == ".xlsx":
            _apply_xlsx(temp_path, changeset)
        else:
            _apply_textual(temp_path, target.suffix.lower(), changeset)
        _validate_output(temp_path, target.suffix.lower())

        if settings.security.create_backups:
            backup_dir = workspace.root / ".pititino" / "backups"
            backup_dir.mkdir(parents=True, exist_ok=True)
            timestamp = datetime.now(UTC).strftime("%Y%m%dT%H%M%S%fZ")
            backup_path = backup_dir / f"{target.name}.{timestamp}.bak"
            suffix = 1
            while backup_path.exists():
                backup_path = backup_dir / f"{target.name}.{timestamp}.{suffix}.bak"
                suffix += 1
            shutil.copy2(target, backup_path)
        os.replace(temp_path, target)
        temp_path = None
        record_transaction(workspace, changeset, "applied", backup_path=backup_path)
        return backup_path or target
    except (OSError, ValueError, ToolExecutionError, UnsupportedFileTypeError) as exc:
        record_transaction(workspace, changeset, "failed", error=str(exc))
        raise ToolExecutionError(f"Unable to apply changes to {target}: {exc}") from exc
    finally:
        if temp_path is not None:
            temp_path.unlink(missing_ok=True)


def _apply_xlsx(path: Path, changeset: ChangeSet) -> None:
    workbook = load_workbook(path)
    try:
        for operation in changeset.operations:
            apply_change_operation(workbook, operation)
        workbook.save(path)
    finally:
        workbook.close()


def _resolve_backup(workspace: Workspace, backup: str) -> Path:
    path = workspace.resolve(backup, must_exist=True)
    backup_dir = workspace.root / ".pititino" / "backups"
    try:
        path.relative_to(backup_dir)
    except ValueError as exc:
        raise UnsupportedFileTypeError("Backup path must be inside .pititino/backups") from exc
    if not path.is_file() or not path.name.endswith(".bak"):
        raise ToolExecutionError("Invalid backup file")
    return path


def _apply_textual(path: Path, suffix: str, changeset: ChangeSet) -> None:
    operation_names = {operation.operation for operation in changeset.operations}
    text_operations = {"text_replace", "text_append"}
    if operation_names & text_operations:
        if not operation_names <= text_operations:
            raise ToolExecutionError("Cannot mix text and structured operations")
        if suffix not in {".txt", ".md"}:
            raise UnsupportedFileTypeError("Text writes require .txt or .md files")
        content = path.read_text(encoding="utf-8")
        for operation in changeset.operations:
            arguments = operation.arguments
            if operation.operation == "text_replace":
                count = arguments["count"]
                content = content.replace(arguments["old"], arguments["new"], count)
            elif operation.operation == "text_append":
                content += arguments["content"]
        path.write_text(content, encoding="utf-8")
        return

    if "structured_set" in operation_names:
        if operation_names != {"structured_set"}:
            raise ToolExecutionError("Cannot mix structured and CSV operations")
        format_name = changeset.operations[0].arguments["format"]
        accepted = {".json"} if format_name == "json" else {".yaml", ".yml"}
        if suffix not in accepted:
            raise UnsupportedFileTypeError(f"{format_name.upper()} writes require matching file extensions")
        parser = json.loads if format_name == "json" else yaml.safe_load
        value = parser(path.read_text(encoding="utf-8"))
        for operation in changeset.operations:
            _set_dotted_value(value, operation.arguments["path"], operation.arguments["value"])
        if format_name == "json":
            path.write_text(json.dumps(value, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
        else:
            path.write_text(yaml.safe_dump(value, sort_keys=False), encoding="utf-8")
        return

    if "csv_append_rows" in operation_names:
        if operation_names != {"csv_append_rows"}:
            raise ToolExecutionError("Cannot mix CSV and other operations")
        if suffix != ".csv":
            raise UnsupportedFileTypeError("CSV writes require a .csv file")
        with path.open("r", encoding="utf-8", newline="") as handle:
            rows = list(csv.reader(handle))
        width = len(rows[0]) if rows else None
        pending_rows = [
            row
            for operation in changeset.operations
            for row in operation.arguments["rows"]
        ]
        if width is not None and any(len(row) != width for row in pending_rows):
            raise ToolExecutionError("CSV row width does not match the header")
        with path.open("a", encoding="utf-8", newline="") as handle:
            writer = csv.writer(handle)
            writer.writerows(pending_rows)
        return

    raise ToolExecutionError("No supported operations for this file type")


def _set_dotted_value(document: object, dotted_path: str, value: object) -> None:
    if not isinstance(document, dict):
        raise ToolExecutionError("Structured root must be an object")
    parts = dotted_path.split(".")
    current = document
    for part in parts[:-1]:
        if not part:
            raise ToolExecutionError("Structured path contains an empty component")
        child = current.get(part)
        if child is None:
            child = {}
            current[part] = child
        if not isinstance(child, dict):
            raise ToolExecutionError(f"Structured path component is not an object: {part}")
        current = child
    if not parts[-1]:
        raise ToolExecutionError("Structured path must not end with a dot")
    current[parts[-1]] = value


def _validate_output(path: Path, suffix: str) -> None:
    if suffix == ".xlsx":
        workbook = load_workbook(path, read_only=True)
        workbook.close()
    elif suffix == ".json":
        json.loads(path.read_text(encoding="utf-8"))
    elif suffix in {".yaml", ".yml"}:
        yaml.safe_load(path.read_text(encoding="utf-8"))
    elif suffix == ".csv":
        with path.open("r", encoding="utf-8", newline="") as handle:
            list(csv.reader(handle))
    elif suffix in {".txt", ".md"}:
        path.read_text(encoding="utf-8")
