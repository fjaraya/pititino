from __future__ import annotations

from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter, range_boundaries
from pydantic import BaseModel, Field

from pititino.config import ExcelConfig
from pititino.errors import UnsupportedFileTypeError, WorkbookReadError
from pititino.transactions.changeset import ChangeOperation, ChangeSet
from pititino.workspace import Workspace


class WorkbookArguments(BaseModel):
    file: str


class SheetArguments(BaseModel):
    file: str
    sheet: str


class RangeArguments(SheetArguments):
    range: str = Field(min_length=1, description="Explicit Excel range such as A1:F100")


class CreateSheetArguments(WorkbookArguments):
    sheet: str = Field(min_length=1, max_length=31)


class WriteRangeArguments(RangeArguments):
    values: list[list[Any]] = Field(min_length=1)


def _workbook_path(workspace: Workspace, file: str) -> Path:
    path = workspace.resolve(file, must_exist=True)
    if path.suffix.lower() != ".xlsx":
        raise UnsupportedFileTypeError("Pititino currently supports .xlsx workbooks, not .xls files")
    if not path.is_file():
        raise FileNotFoundError(path)
    return path


def inspect_workbook(
    workspace_or_path: Workspace | str | Path, arguments: WorkbookArguments | None = None
) -> dict[str, Any]:
    """Return bounded workbook metadata, accepting the old path-only API for compatibility."""
    if isinstance(workspace_or_path, Workspace):
        assert arguments is not None
        path = _workbook_path(workspace_or_path, arguments.file)
        relative_path = str(path.relative_to(workspace_or_path.root))
    else:
        path = Path(workspace_or_path).expanduser().resolve()
        relative_path = str(path)
        if path.suffix.lower() != ".xlsx":
            raise UnsupportedFileTypeError("Pititino currently supports .xlsx workbooks, not .xls files")
        if not path.is_file():
            raise FileNotFoundError(path)

    try:
        workbook = load_workbook(path, read_only=True, data_only=False)
    except (OSError, ValueError) as exc:
        raise WorkbookReadError(f"Unable to open workbook {path}: {exc}") from exc
    try:
        return {
            "filename": path.name,
            "path": relative_path,
            "sheets": [
                {
                    "name": sheet.title,
                    "rows": sheet.max_row,
                    "columns": sheet.max_column,
                    "state": sheet.sheet_state,
                }
                for sheet in workbook.worksheets
            ],
        }
    finally:
        workbook.close()


def inspect_sheet(workspace: Workspace, arguments: SheetArguments, config: ExcelConfig) -> dict[str, Any]:
    path = _workbook_path(workspace, arguments.file)
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        if arguments.sheet not in workbook.sheetnames:
            raise WorkbookReadError(f"Worksheet not found: {arguments.sheet}")
        sheet = workbook[arguments.sheet]
        rows = list(sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 10), values_only=True))
        header_index = next((index for index, row in enumerate(rows) if any(value is not None for value in row)), None)
        headers = list(rows[header_index]) if header_index is not None else []
        samples = rows[header_index + 1 : header_index + 6] if header_index is not None else rows[:5]
        types = Counter(type(value).__name__ for row in samples for value in row if value is not None)
        return {
            "file": str(path.relative_to(workspace.root)),
            "sheet": sheet.title,
            "rows": sheet.max_row,
            "columns": sheet.max_column,
            "state": sheet.sheet_state,
            "header_row": header_index + 1 if header_index is not None else None,
            "headers": headers,
            "sample_rows": [list(row) for row in samples],
            "sample_types": dict(types),
            "sample_limit": min(config.max_rows_per_read, 5),
        }
    finally:
        workbook.close()


def read_range(workspace: Workspace, arguments: RangeArguments, config: ExcelConfig) -> dict[str, Any]:
    path = _workbook_path(workspace, arguments.file)
    try:
        min_col, min_row, max_col, max_row = range_boundaries(arguments.range)
    except ValueError as exc:
        raise WorkbookReadError(f"Invalid Excel range: {arguments.range}") from exc
    row_count = max_row - min_row + 1
    cell_count = row_count * (max_col - min_col + 1)
    if row_count > config.max_rows_per_read:
        raise WorkbookReadError(f"Range exceeds max_rows_per_read ({config.max_rows_per_read})")
    if cell_count > config.max_cells_per_read:
        raise WorkbookReadError(f"Range exceeds max_cells_per_read ({config.max_cells_per_read})")

    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        if arguments.sheet not in workbook.sheetnames:
            raise WorkbookReadError(f"Worksheet not found: {arguments.sheet}")
        sheet = workbook[arguments.sheet]
        values = [
            list(row)
            for row in sheet.iter_rows(
                min_row=min_row,
                max_row=max_row,
                min_col=min_col,
                max_col=max_col,
                values_only=True,
            )
        ]
        return {
            "file": str(path.relative_to(workspace.root)),
            "sheet": sheet.title,
            "range": f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}",
            "rows": row_count,
            "columns": max_col - min_col + 1,
            "values": values,
        }
    finally:
        workbook.close()


def propose_create_sheet(arguments: CreateSheetArguments) -> ChangeSet:
    return ChangeSet(
        target=arguments.file,
        summary=f'Create worksheet "{arguments.sheet}"',
        operations=[
            ChangeOperation(
                operation="create_sheet",
                description=f'Create worksheet "{arguments.sheet}"',
                arguments={"sheet": arguments.sheet},
            )
        ],
    )


def propose_write_range(arguments: WriteRangeArguments) -> ChangeSet:
    return ChangeSet(
        target=arguments.file,
        summary=f"Write {arguments.sheet}!{arguments.range}",
        operations=[
            ChangeOperation(
                operation="write_range",
                description=f"Write {arguments.sheet}!{arguments.range}",
                arguments={
                    "sheet": arguments.sheet,
                    "range": arguments.range,
                    "values": arguments.values,
                },
            )
        ],
    )


def apply_change_operation(workbook: Any, operation: ChangeOperation) -> None:
    if operation.operation == "create_sheet":
        sheet_name = operation.arguments["sheet"]
        if sheet_name in workbook.sheetnames:
            raise WorkbookReadError(f"Worksheet already exists: {sheet_name}")
        workbook.create_sheet(sheet_name)
        return

    if operation.operation == "write_range":
        sheet_name = operation.arguments["sheet"]
        if sheet_name not in workbook.sheetnames:
            raise WorkbookReadError(f"Worksheet not found: {sheet_name}")
        min_col, min_row, max_col, max_row = range_boundaries(operation.arguments["range"])
        values = operation.arguments["values"]
        if len(values) != max_row - min_row + 1 or any(
            len(row) != max_col - min_col + 1 for row in values
        ):
            raise WorkbookReadError("Write values do not match the requested range dimensions")
        sheet = workbook[sheet_name]
        for row_offset, row in enumerate(values):
            for col_offset, value in enumerate(row):
                sheet.cell(row=min_row + row_offset, column=min_col + col_offset, value=value)
        return

    raise WorkbookReadError(f"Unsupported workbook operation: {operation.operation}")
