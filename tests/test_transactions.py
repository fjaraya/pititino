from openpyxl import Workbook, load_workbook

from pititino.config import Settings
from pititino.tools.excel import (
    CreateSheetArguments,
    WriteRangeArguments,
    propose_create_sheet,
    propose_write_range,
)
from pititino.transactions.executor import apply_changeset
from pititino.workspace import Workspace


def test_apply_changeset_creates_backup_and_reopens(tmp_path) -> None:
    path = tmp_path / "sales.xlsx"
    workbook = Workbook()
    workbook.active.title = "Sales"
    workbook.save(path)

    changes = propose_create_sheet(CreateSheetArguments(file="sales.xlsx", sheet="Overview"))
    result = apply_changeset(changes, Workspace(tmp_path), Settings())

    assert result.name.endswith(".bak")
    assert result.exists()
    output = load_workbook(path, read_only=True)
    assert output.sheetnames == ["Sales", "Overview"]
    output.close()


def test_apply_changeset_writes_range(tmp_path) -> None:
    path = tmp_path / "sales.xlsx"
    workbook = Workbook()
    workbook.active.title = "Sales"
    workbook.save(path)

    changes = propose_write_range(
        WriteRangeArguments(
            file="sales.xlsx", sheet="Sales", range="A1:B2", values=[["A", "B"], [1, 2]]
        )
    )
    apply_changeset(changes, Workspace(tmp_path), Settings())

    output = load_workbook(path, data_only=True)
    assert list(output["Sales"].values) == [("A", "B"), (1, 2)]
    output.close()
