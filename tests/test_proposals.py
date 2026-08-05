from openpyxl import Workbook, load_workbook

from pititino.config import Settings
from pititino.tools import build_registry
from pititino.transactions.changeset import ChangeSet
from pititino.workspace import Workspace


def test_mutating_tool_returns_a_proposal_without_modifying_file(tmp_path) -> None:
    path = tmp_path / "sales.xlsx"
    workbook = Workbook()
    workbook.active.title = "Sales"
    workbook.save(path)

    registry = build_registry(Workspace(tmp_path), Settings())
    result = registry.invoke("excel.create_sheet", {"file": "sales.xlsx", "sheet": "Overview"})

    assert isinstance(result, ChangeSet)
    assert result.requires_confirmation is True
    output = load_workbook(path, read_only=True)
    assert output.sheetnames == ["Sales"]
    output.close()
