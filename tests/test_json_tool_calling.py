from types import SimpleNamespace

import pytest
from openpyxl import Workbook

from pititino.agent.runtime import AgentRuntime
from pititino.config import Settings
from pititino.errors import ModelEndpointError
from pititino.tools import build_registry
from pititino.workspace import Workspace


def response(content):
    return SimpleNamespace(
        choices=[SimpleNamespace(message=SimpleNamespace(content=content, tool_calls=[]))]
    )


class JsonClient:
    def __init__(self):
        self.calls = []

    async def complete(self, messages, tools):
        self.calls.append((messages, tools))
        if len(self.calls) == 1:
            return response('{"action":"filesystem.list","arguments":{"path":"."}}')
        return response("The workspace was inspected.")


class AutoFallbackClient(JsonClient):
    async def complete(self, messages, tools):
        self.calls.append((messages, tools))
        if len(self.calls) == 1:
            raise ModelEndpointError("tools are unsupported")
        if len(self.calls) == 2:
            return response('{"action":"filesystem.list","arguments":{"path":"."}}')
        return response("Fallback inspection completed.")


class SilentNativeClient(JsonClient):
    async def complete(self, messages, tools):
        self.calls.append((messages, tools))
        if len(self.calls) == 1:
            return response("I will inspect the workspace.")
        if len(self.calls) == 2:
            return response('{"action":"filesystem.list","arguments":{"path":"."}}')
        return response("Inspection completed after fallback.")


class JsonExcelClient:
    def __init__(self):
        self.calls = []

    async def complete(self, messages, tools):
        self.calls.append((messages, tools))
        if len(self.calls) == 1:
            return response(
                '{"action":"excel.inspect_workbook","arguments":{"file":"book.xlsx"}}'
            )
        if len(self.calls) == 2:
            return response(
                '{"action":"excel.create_sheet","arguments":{"file":"book.xlsx","sheet":"Overview"}}'
            )
        return response("The overview sheet is ready for approval.")


class EmptyAfterToolClient:
    def __init__(self):
        self.calls = []

    async def complete(self, messages, tools):
        self.calls.append((messages, tools))
        if len(self.calls) == 1:
            return response('{"action":"filesystem.list","arguments":{"path":"."}}')
        if len(self.calls) == 2:
            return response("")
        return response("The workspace was inspected.")


@pytest.mark.anyio
async def test_json_mode_executes_validated_actions_without_native_tools(tmp_path) -> None:
    client = JsonClient()
    settings = Settings(model={"tool_calling": "json"})
    runtime = AgentRuntime(settings, build_registry(Workspace(tmp_path), settings), client)

    result = await runtime.run("List the workspace")

    assert result == "The workspace was inspected."
    assert all(tools == [] for _, tools in client.calls)


@pytest.mark.anyio
async def test_auto_mode_falls_back_to_json_actions(tmp_path) -> None:
    client = AutoFallbackClient()
    settings = Settings(model={"tool_calling": "auto"})
    runtime = AgentRuntime(settings, build_registry(Workspace(tmp_path), settings), client)

    result = await runtime.run("List the workspace")

    assert result == "Fallback inspection completed."
    assert client.calls[0][1]
    assert client.calls[1][1] == []


@pytest.mark.anyio
async def test_auto_mode_falls_back_when_native_response_has_no_tool_call(tmp_path) -> None:
    client = SilentNativeClient()
    settings = Settings(model={"tool_calling": "auto"})
    activities = []
    runtime = AgentRuntime(
        settings,
        build_registry(Workspace(tmp_path), settings),
        client,
        on_tool_activity=activities.append,
    )

    result = await runtime.run("List the workspace")

    assert result == "Inspection completed after fallback."
    assert client.calls[0][1]
    assert client.calls[1][1] == []
    assert any("native response contained no tool call" in activity for activity in activities)


@pytest.mark.anyio
async def test_json_mode_continues_from_workbook_inspection_to_mutation(tmp_path) -> None:
    workbook = Workbook()
    workbook.save(tmp_path / "book.xlsx")
    client = JsonExcelClient()
    settings = Settings(model={"tool_calling": "json"})
    runtime = AgentRuntime(settings, build_registry(Workspace(tmp_path), settings), client)

    result = await runtime.run("Inspect book.xlsx and create an Overview sheet")

    assert result == "The overview sheet is ready for approval."
    assert len(runtime.pending_changes) == 1
    assert runtime.pending_changes[0].target == "book.xlsx"
    assert runtime.pending_changes[0].operations[0].operation == "create_sheet"
    assert client.calls[1][0][-2]["role"] == "assistant"
    assert client.calls[1][0][-1]["role"] == "user"
    assert all(message["role"] != "tool" for message in client.calls[1][0])


@pytest.mark.anyio
async def test_empty_json_response_is_retried_after_tool_execution(tmp_path) -> None:
    client = EmptyAfterToolClient()
    settings = Settings(model={"tool_calling": "json"})
    runtime = AgentRuntime(settings, build_registry(Workspace(tmp_path), settings), client)

    result = await runtime.run("List the workspace")

    assert result == "The workspace was inspected."
    assert len(client.calls) == 3
